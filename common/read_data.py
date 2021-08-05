#coding:utf-8
import openpyxl
from get_path import *
import json
import pymysql
import yaml


def read_excel(file_path,sheet_name):
    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_name]
    data_list = []
    row = sheet.max_row
    # col = sheet.max_column
    for i in range(2,row+1):
        module = sheet.cell(i,1).value
        testcase = sheet.cell(i, 2).value
        descript = sheet.cell(i, 3).value
        method = sheet.cell(i, 4).value
        code = sheet.cell(i, 5).value
        pre_sql = sheet.cell(i, 6).value
        url = sheet.cell(i, 7).value
        headers = sheet.cell(i, 8).value
        body = sheet.cell(i, 9).value
        assert_type = sheet.cell(i, 10).value
        assert_data = sheet.cell(i, 11).value

        data_dict = {"module":module,"code":code,"testcase":testcase,"descript":descript,"method":method,"url":url,"headers":headers,
                     "body":body,"assert_data":assert_data,"pre_sql":pre_sql,"assert_type":assert_type}
        data_list.append(data_dict)
    return data_list

def write_into_excel(file_path,sheet_name,row,cols,value):
    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_name]
    sheet.cell(row,cols).value=value
    book.save(file_path)

def read_const_excel(file_path,sheet_name):
    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_name]
    # data_list = []
    # row = sheet.max_row
    environment = sheet.cell(2,1).value
    session = sheet.cell(2,2).value
    account = sheet.cell(2,3).value
    merchantId = sheet.cell(2, 4).value
    merchantId = merchantId.replace("\n","")
    operatorId = sheet.cell(2, 5).value
    data_dict = {"environment":environment,"session":session,"account":account,"merchantId":merchantId,"operatorId":operatorId}
    return data_dict

def get_yaml_data(file_path):
    # menu = "商户管理"
    with open(file_path,"r",encoding="utf-8") as f:
        data = yaml.load(f,Loader=yaml.FullLoader)
    # print(data)
    # print(data["menu_element"]%menu)
    return data


def con_mysql(flag,env,db,key,sql):
    """
    flag:数据标识
    env:环境，sit或dev
    key:selectall：返回全部
    key:selectone:返回一条
    key:update更新
    key:insert 插入
    key：delete或者del:删除
    sql:list，可执行多条

    """
    # print(db_config)
    data = get_yaml_data(db_config)[flag]
    # env = data["default"]
    db_data = data[env]
    # print("数据库环境是：",env)
    host = db_data["host"]
    username = db_data["username"]
    password = db_data["password"]
    # db = db_data["db"]
    port = db_data["port"]
    con = pymysql.connect(host=host,user=username,passwd=password,db=db,port=port)
    #pymysql.cursors.DictCursor让返回的数据为key:value形式
    cur = con.cursor(pymysql.cursors.DictCursor)
    db_sql = []
    for sq in sql:
        cur.execute(sq)
        if key=="selectall":
            db_sql.append(cur.fetchall())
        elif key=="selectone":
            db_sql.append(cur.fetchall())
        elif key in ("update","insert","delete","del"):
            con.commit()
        else:print("key不对")
    return db_sql
    cur.close()
    con.close()









if __name__=="__main__":
    sql = ["select id role_id from merchant_role where merchant_id='1948893113663488001';","select id from merchant_user_role_relate where merchant_id='1948893113663488001';"]
    # data = read_const_excel(iflying_manage_data,"const")
    # print(data)
    # print(headers)
    # print(type(headers))
    # write_into_excel(iflying_data,"api")
    res = con_mysql("merchant_center","sit","merchant","selectone",sql)
    print(list(res[0][0].values())[0])
    print(list(res[1][0].values())[0])
